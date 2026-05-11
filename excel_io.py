"""
Excel export/import for Nitya VFX Studio.
Uses openpyxl — no server, fully offline.
"""
import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ── Status color map (Excel ARGB hex) ────────────────────────────────────────
STATUS_FILLS = {
    "Pending":  ("FF1A1A1A", "FFE67E22"),   # (bg, fg)
    "WIP":      ("FF001830", "FF4A90D9"),
    "Review":   ("FF150020", "FF9B59B6"),
    "Approved": ("FF002010", "FF2ECC71"),
    "Hold":     ("FF200000", "FFE74C3C"),
    "Retake":   ("FF1A1000", "FFF39C12"),
    "N/A":      ("FF1A1A1A", "FF6C7A89"),
}

PRIORITY_FILLS = {
    "Low":      "FF6C7A89",
    "Normal":   "FF4A90D9",
    "High":     "FFE67E22",
    "Critical": "FFE74C3C",
}

THIN = Side(style='thin', color="FF2D3561")
MED  = Side(style='medium', color="FFE94560")

HEADER_FONT  = Font(name='Calibri', bold=True, color="FFFFFFFF", size=10)
HEADER_FILL  = PatternFill("solid", fgColor="FF0F3460")
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
HEADER_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=MED)

DATA_FONT    = Font(name='Calibri', size=9)
DATA_ALIGN   = Alignment(vertical='center', wrap_text=False)
DATA_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT   = Font(name='Calibri', bold=True, size=14, color="FFE94560")
META_FONT    = Font(name='Calibri', italic=True, size=9, color="FF6C7A89")


COLUMNS = [
    ("Project Name",  20),
    ("Sequence",      14),
    ("Shot Name",     18),
    ("Artist",        18),
    ("Frame Count",   12),
    ("Start Frame",   12),
    ("End Frame",     12),
    ("ETA",           14),
    ("Status",        14),
    ("Priority",      12),
    ("Notes",         40),
]


def export_to_excel(project_info: dict, shots: list, filepath: str):
    """
    Export all shots to a professionally formatted Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shots"

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"NITYA VFX STUDIO — {project_info.get('display_name','Project').upper()}"
    title_cell.font = TITLE_FONT
    title_cell.fill = PatternFill("solid", fgColor="FF16213E")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:K2")
    meta_cell = ws["A2"]
    meta_cell.value = (
        f"Type: {project_info.get('project_type','')}   |   "
        f"Client: {project_info.get('client','')}   |   "
        f"Exported: {datetime.date.today().strftime('%d-%b-%Y')}   |   "
        f"Total Shots: {len(shots)}"
    )
    meta_cell.font = META_FONT
    meta_cell.fill = PatternFill("solid", fgColor="FF0F3460")
    meta_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Column headers (row 3) ───────────────────────────────────────────────
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = col_name.upper()
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[3].height = 24

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, shot in enumerate(shots, start=4):
        status = shot.get("status", "Pending")
        priority = shot.get("priority", "Normal")
        bg_hex, fg_hex = STATUS_FILLS.get(status, ("FF1A1A2E", "FFE0E0E0"))

        row_data = [
            project_info.get("display_name", ""),
            shot.get("sequence", ""),
            shot.get("shot_name", ""),
            shot.get("artist", ""),
            shot.get("frame_count", 0),
            shot.get("start_frame", 1001),
            shot.get("end_frame", 1001),
            shot.get("eta", ""),
            status,
            priority,
            shot.get("notes", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(name='Calibri', size=9, color=fg_hex)
            cell.border = DATA_BORDER
            cell.alignment = DATA_ALIGN

            # Row background from status
            if col_idx == 9:  # Status column — special highlight
                cell.fill = PatternFill("solid", fgColor=bg_hex[2:])  # strip FF prefix
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 10:  # Priority column
                p_color = PRIORITY_FILLS.get(priority, "FF4A90D9")
                cell.fill = PatternFill("solid", fgColor=p_color[2:])
                cell.font = Font(name='Calibri', size=9, color="FFFFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.fill = PatternFill("solid", fgColor="FF141428" if row_idx % 2 == 0 else "FF1A1A38")

        ws.row_dimensions[row_idx].height = 18

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter ──────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:K{max(3, 3 + len(shots))}"

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2["A1"].value = "STATUS SUMMARY"
    ws2["A1"].font = TITLE_FONT

    status_counts = {}
    for shot in shots:
        st = shot.get("status", "Pending")
        status_counts[st] = status_counts.get(st, 0) + 1

    artist_counts = {}
    for shot in shots:
        ar = shot.get("artist", "Unassigned") or "Unassigned"
        artist_counts[ar] = artist_counts.get(ar, 0) + 1

    ws2["A2"].value = "Status"
    ws2["A2"].font = HEADER_FONT
    ws2["A2"].fill = HEADER_FILL
    ws2["B2"].value = "Count"
    ws2["B2"].font = HEADER_FONT
    ws2["B2"].fill = HEADER_FILL

    r = 3
    for st, cnt in sorted(status_counts.items()):
        ws2.cell(row=r, column=1).value = st
        ws2.cell(row=r, column=2).value = cnt
        bg_hex, fg_hex = STATUS_FILLS.get(st, ("FF1A1A2E", "FFE0E0E0"))
        ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=bg_hex[2:])
        ws2.cell(row=r, column=1).font = Font(color=fg_hex, bold=True)
        r += 1

    r += 1
    ws2.cell(row=r, column=1).value = "Artist Workload"
    ws2.cell(row=r, column=1).font = TITLE_FONT
    r += 1
    ws2.cell(row=r, column=1).value = "Artist"
    ws2.cell(row=r, column=1).font = HEADER_FONT
    ws2.cell(row=r, column=1).fill = HEADER_FILL
    ws2.cell(row=r, column=2).value = "Shot Count"
    ws2.cell(row=r, column=2).font = HEADER_FONT
    ws2.cell(row=r, column=2).fill = HEADER_FILL
    r += 1
    for ar, cnt in sorted(artist_counts.items(), key=lambda x: -x[1]):
        ws2.cell(row=r, column=1).value = ar
        ws2.cell(row=r, column=2).value = cnt
        ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor="FF141428")
        r += 1

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 14

    # ── Page setup ───────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "3:3"

    wb.save(filepath)
    return filepath




def import_from_excel(filepath: str) -> list:
    """
    Parse an Excel file and return a list of shot dicts.
    Supports both our export format and simple tabular formats.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    shots = []
    header_row = None
    header_map = {}

    COLUMN_ALIASES = {
        "project name": "project_name",
        "sequence": "sequence",
        "shot name": "shot_name",
        "shot": "shot_name",
        "artist": "artist",
        "frame count": "frame_count",
        "frames": "frame_count",
        "start frame": "start_frame",
        "end frame": "end_frame",
        "eta": "eta",
        "status": "status",
        "priority": "priority",
        "notes": "notes",
        "note": "notes",
    }

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > 5 and not header_map:
            break  # give up looking for header

        if all(c is None for c in row):
            continue

        # Detect header row
        if not header_map:
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                key = str(cell).strip().lower()
                if key in COLUMN_ALIASES:
                    header_map[col_idx] = COLUMN_ALIASES[key]
            if header_map:
                header_row = row_idx
            continue

        # Data row
        if row_idx > header_row:
            shot = {}
            for col_idx, field in header_map.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    shot[field] = str(val).strip() if val is not None else ""
            if shot.get("shot_name"):
                shots.append(shot)

    wb.close()
    return shots
